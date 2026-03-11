from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

import monitor_front_basket as monitor


@dataclass(frozen=True)
class CsvSource:
    name: str
    xlsx_column: str
    path: Path


def _default_output_path(src: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return src.with_name(f"{src.stem} (updated) {stamp}{src.suffix}")


def _pick_latest_xlsx(directory: Path) -> Path | None:
    files = [p for p in directory.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _norm_barcode(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip().replace("\t", "")
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits or None


def _load_config(path: Path) -> list[CsvSource]:
    data = json.loads(path.read_text(encoding="utf-8"))
    sources: list[CsvSource] = []
    for src in data.get("sources", []):
        if src.get("type") != "csv":
            continue
        sources.append(
            CsvSource(
                name=str(src.get("name") or src.get("xlsx_column") or "csv"),
                xlsx_column=str(src["xlsx_column"]),
                path=(path.parent / str(src["path"])).resolve(),
            )
        )
    if not sources:
        raise SystemExit(f"В конфиге нет источников type=csv: {path}")
    return sources


def _read_csv_prices(path: Path) -> dict[str, float]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    prices: dict[str, float] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return prices
        for row in reader:
            bc = _norm_barcode(row.get("barcode"))
            if not bc:
                continue
            price = monitor._to_number(row.get("price"))
            if price is None or price <= 0:
                continue
            prices[bc] = float(price)
    return prices


def _write_demo_csv(path: Path, rows: list[tuple[str, float]]) -> None:
    path.write_text("", encoding="utf-8")
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["barcode", "price"])
        writer.writeheader()
        for bc, price in rows:
            writer.writerow({"barcode": bc, "price": round(price, 2)})


def update_xlsx(
    src_xlsx: Path,
    *,
    config_path: Path,
    sheet_name: str | None,
    barcode_header: str,
    our_price_header: str,
    threshold: float,
    demo_generate: bool,
) -> tuple[Path, dict[str, int], dict[str, int]]:
    sources = _load_config(config_path)

    wb = openpyxl.load_workbook(src_xlsx)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = monitor._find_header_row(ws)
    col_map = monitor._build_column_map(ws, header_row)

    barcode_key = monitor._norm_header(barcode_header)
    if barcode_key not in col_map:
        raise SystemExit(f'Не нашёл столбец "{barcode_header}" для ключа (штрих-код).')
    barcode_col = col_map[barcode_key].col_idx

    our_key = monitor._norm_header(our_price_header)
    if our_key not in col_map:
        raise SystemExit(f'Не нашёл столбец нашей цены "{our_price_header}".')
    our_col = col_map[our_key].col_idx

    missing_cols: list[str] = []
    source_cols: dict[str, int] = {}
    for s in sources:
        key = monitor._norm_header(s.xlsx_column)
        if key in col_map:
            source_cols[s.xlsx_column] = col_map[key].col_idx
        else:
            missing_cols.append(s.xlsx_column)
    if missing_cols:
        raise SystemExit(f"Не нашёл в XLSX колонки конкурентов: {', '.join(missing_cols)}")

    loaded_prices: dict[str, dict[str, float]] = {}
    for s in sources:
        if demo_generate and not s.path.exists():
            demo_rows: list[tuple[str, float]] = []
            for r in range(header_row + 1, ws.max_row + 1):
                bc = _norm_barcode(ws.cell(r, barcode_col).value)
                our_price = monitor._to_number(ws.cell(r, our_col).value)
                if not bc or our_price is None or our_price <= 0:
                    continue
                demo_rows.append((bc, float(our_price) * 1.15))
                if len(demo_rows) >= 200:
                    break
            _write_demo_csv(s.path, demo_rows)

        loaded_prices[s.xlsx_column] = _read_csv_prices(s.path)

    updated_counts: dict[str, int] = {s.xlsx_column: 0 for s in sources}
    for r in range(header_row + 1, ws.max_row + 1):
        bc = _norm_barcode(ws.cell(r, barcode_col).value)
        if not bc:
            continue
        for s in sources:
            prices = loaded_prices[s.xlsx_column]
            if bc not in prices:
                continue
            ws.cell(r, source_cols[s.xlsx_column]).value = round(prices[bc], 2)
            updated_counts[s.xlsx_column] += 1

    alert_stats = monitor.apply_price_index(
        ws,
        our_price_header=our_price_header,
        competitor_headers=tuple(s.xlsx_column for s in sources),
        threshold=threshold,
    )

    out_path = _default_output_path(src_xlsx)
    wb.save(out_path)
    return out_path, updated_counts, alert_stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Обновляет цены конкурентов (по конфигу) в новой копии XLSX и считает PI (конкурент / наша)."
    )
    parser.add_argument("--file", type=Path, default=None, help="Путь к исходному .xlsx (по умолчанию самый новый в папке).")
    parser.add_argument("--config", type=Path, default=Path("competitors.json"), help="Конфиг источников цен конкурентов (json).")
    parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый).")
    parser.add_argument("--barcode", type=str, default="Штрих-код", help='Заголовок колонки-ключа (по умолчанию "Штрих-код").')
    parser.add_argument("--our", type=str, default="наша цена", help='Заголовок колонки нашей цены (по умолчанию "наша цена").')
    parser.add_argument("--threshold", type=float, default=1.1, help="Порог PI (симметрично вверх/вниз). 1.1 = 10%.")
    parser.add_argument("--demo-generate", action="store_true", help="Если CSV нет — сгенерировать демо-цены (наша*1.15) для теста.")
    args = parser.parse_args()

    if args.file is None:
        picked = _pick_latest_xlsx(Path.cwd())
        if picked is None:
            raise SystemExit("В текущей директории не найдено ни одного .xlsx файла.")
        src_path = picked
    else:
        src_path = args.file

    out_path, updated_counts, alert_stats = update_xlsx(
        src_path,
        config_path=args.config,
        sheet_name=args.sheet,
        barcode_header=args.barcode,
        our_price_header=args.our,
        threshold=float(args.threshold),
        demo_generate=bool(args.demo_generate),
    )

    print(f"OK: {out_path.name}")
    for k, v in updated_counts.items():
        print(f"- updated {k}: {v}")
    for k, v in alert_stats.items():
        print(f"- alerts {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

