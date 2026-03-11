from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill


DEFAULT_COMPETITOR_COLUMNS: tuple[str, ...] = (
    "Розница",
    "Розница Х5",
    "Розница Магнит",
    "Розница Монетка",
    "Розница Лента",
)


ALERT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


@dataclass(frozen=True)
class ColumnRef:
    header: str
    col_idx: int


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or s == "-":
            return None
        s = s.replace("\u00a0", "").replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_scan_rows: int = 15) -> int:
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        norms = {_norm_header(v) for v in row if v is not None}
        if "артикул" in norms and ("наша цена" in norms or "нашацена" in norms):
            return r
    return 1


def _build_column_map(
    ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int
) -> dict[str, ColumnRef]:
    col_map: dict[str, ColumnRef] = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(header_row, c).value
        key = _norm_header(header)
        if not key:
            continue
        if key not in col_map:
            col_map[key] = ColumnRef(header=str(header).strip(), col_idx=c)
    return col_map


def _next_available_col(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> int:
    last = ws.max_column
    while last > 1 and ws.cell(header_row, last).value in (None, ""):
        last -= 1
    return last + 1


def _ratio_alert(price_a: float, price_b: float, threshold: float) -> bool:
    if price_a <= 0 or price_b <= 0:
        return False
    ratio = price_a / price_b
    return ratio > threshold or (1 / ratio) > threshold


def _default_output_path(src: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return src.with_name(f"{src.stem} (monitor) {stamp}{src.suffix}")


def _pick_latest_xlsx(directory: Path) -> Path | None:
    files = [p for p in directory.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def monitor(
    src_path: Path,
    *,
    sheet_name: str | None,
    our_price_header: str,
    competitor_headers: tuple[str, ...],
    threshold: float,
) -> tuple[Path, dict[str, int]]:
    wb = openpyxl.load_workbook(src_path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    stats = apply_price_index(
        ws,
        our_price_header=our_price_header,
        competitor_headers=competitor_headers,
        threshold=threshold,
    )

    out_path = _default_output_path(src_path)
    wb.save(out_path)
    return out_path, stats


def apply_price_index(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    our_price_header: str,
    competitor_headers: tuple[str, ...],
    threshold: float,
) -> dict[str, int]:
    header_row = _find_header_row(ws)
    col_map = _build_column_map(ws, header_row)

    our_key = _norm_header(our_price_header)
    if our_key not in col_map:
        raise SystemExit(
            f'Не нашёл столбец нашей цены "{our_price_header}". '
            f"Доступные заголовки: {', '.join(sorted({v.header for v in col_map.values()}))}"
        )
    our_col = col_map[our_key].col_idx

    competitor_cols: list[ColumnRef] = []
    missing: list[str] = []
    for h in competitor_headers:
        key = _norm_header(h)
        if key in col_map:
            competitor_cols.append(col_map[key])
        else:
            missing.append(h)

    if missing:
        raise SystemExit(f"Не нашёл столбцы конкурентов: {', '.join(missing)}")

    pi_cols: dict[str, int] = {}
    insert_col = _next_available_col(ws, header_row)
    for comp in competitor_cols:
        pi_header = f"PI {comp.header}"
        existing = col_map.get(_norm_header(pi_header))
        if existing is not None:
            pi_cols[comp.header] = existing.col_idx
        else:
            ws.cell(header_row, insert_col).value = pi_header
            pi_cols[comp.header] = insert_col
            insert_col += 1

    stats: dict[str, int] = {c.header: 0 for c in competitor_cols}

    for r in range(header_row + 1, ws.max_row + 1):
        our_price = _to_number(ws.cell(r, our_col).value)
        if our_price is None or our_price <= 0:
            continue

        for comp in competitor_cols:
            comp_cell = ws.cell(r, comp.col_idx)
            comp_price = _to_number(comp_cell.value)
            pi_cell = ws.cell(r, pi_cols[comp.header])

            if comp_price is None or comp_price <= 0:
                pi_cell.value = None
                continue

            pi = comp_price / our_price
            pi_cell.value = round(pi, 4)

            if _ratio_alert(comp_price, our_price, threshold):
                stats[comp.header] += 1
                comp_cell.fill = ALERT_FILL
                pi_cell.fill = ALERT_FILL

    return stats




def main() -> int:
    parser = argparse.ArgumentParser(
        description="Мониторинг Front basket: считает price index и помечает отклонения, не трогая исходный XLSX."
    )
    parser.add_argument("--file", type=Path, default=None, help="Путь к исходному .xlsx (по умолчанию берётся самый новый в папке).")
    parser.add_argument("--sheet", type=str, default=None, help="Имя листа (по умолчанию первый).")
    parser.add_argument("--our", type=str, default="наша цена", help='Заголовок столбца с нашей ценой (по умолчанию "наша цена").')
    parser.add_argument(
        "--competitors",
        type=str,
        default=";".join(DEFAULT_COMPETITOR_COLUMNS),
        help='Список заголовков конкурентов через ";" (по умолчанию стандартный набор).',
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=1.1,
        help="Порог отклонения индекса (симметрично вверх/вниз). 1.1 = 10%.",
    )
    args = parser.parse_args()

    src_path: Path
    if args.file is None:
        picked = _pick_latest_xlsx(Path.cwd())
        if picked is None:
            raise SystemExit("В текущей директории не найдено ни одного .xlsx файла.")
        src_path = picked
    else:
        src_path = args.file

    competitors = tuple(h.strip() for h in str(args.competitors).split(";") if h.strip())
    out_path, stats = monitor(
        src_path,
        sheet_name=args.sheet,
        our_price_header=args.our,
        competitor_headers=competitors,
        threshold=float(args.threshold),
    )

    print(f"OK: {out_path.name}")
    for k, v in stats.items():
        print(f"- {k}: alerts={v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
